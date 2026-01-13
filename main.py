#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Veo3Free - AI生成工具 PyWebview + React 版本
"""

import asyncio
import json
import os
import sys
import base64
import io
import subprocess
import platform
import threading
import webbrowser
from datetime import datetime
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler
from functools import partial

# Windows 下设置输出编码为 UTF-8
if sys.platform == 'win32':
    if sys.stdout is not None and hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if sys.stderr is not None and hasattr(sys.stderr, 'buffer'):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

    # 强制 pywebview 使用 WinRT 后端（避免 pythonnet/WinForms 依赖）
    os.environ['PYWEBVIEW_BACKEND'] = 'winrt'

from PIL import Image

try:
    from loguru import logger
except ImportError:
    print("请安装 loguru: pip install loguru")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("请安装 openpyxl: pip install openpyxl")
    Workbook = None
    load_workbook = None

try:
    from websockets.server import serve
except ImportError:
    print("请安装 websockets: pip install websockets")
    sys.exit(1)

try:
    import webview
except ImportError:
    print("请安装 pywebview: pip install pywebview")
    sys.exit(1)

from version import get_version
from updater import check_for_updates, open_download_page

# 确定输出目录位置
if getattr(sys, 'frozen', False):
    # 打包后，使用用户文档目录
    OUTPUT_DIR = Path.home() / "Documents" / "veo3free" / "output"
else:
    # 开发模式，使用项目目录
    OUTPUT_DIR = Path("output")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 配置loguru日志
LOGS_DIR = OUTPUT_DIR.parent / "logs" if getattr(sys, 'frozen', False) else Path("logs")
LOGS_DIR.mkdir(parents=True, exist_ok=True)

# 获取版本号用于日志
from version import __version__ as APP_VERSION

logger.remove()  # 移除默认handler

# 日志格式：时间 | 版本 | 文件:行号 | 消息
LOG_FORMAT_CONSOLE = "<green>{time:HH:mm:ss}</green> | <cyan>v{extra[ver]}</cyan> | <level>{message}</level>"
LOG_FORMAT_FILE = "{time:YYYY-MM-DD HH:mm:ss} | v{extra[ver]} | {name}:{line} | {level: <8} | {message}"

# 控制台输出
logger.add(
    lambda msg: print(msg, end=""),
    format=LOG_FORMAT_CONSOLE,
    level="INFO",
    colorize=True,
    filter=lambda record: not record["extra"].get("file_only")
)

# 文件日志（完整格式，包含代码位置）
log_file = LOGS_DIR / "veo3free.log"
logger.add(
    log_file,
    format=LOG_FORMAT_FILE,
    rotation="10 MB",
    retention="7 days",
    encoding="utf-8"
)

# 绑定版本号到所有日志
logger = logger.bind(ver=APP_VERSION)


def get_logger():
    """获取绑定了版本号的 logger"""
    return logger.bind(ver=APP_VERSION)


def log_error_to_file(message: str, exception: Exception = None):
    """记录错误：控制台显示ASCII安全版本，文件记录完整信息"""
    if exception:
        # 控制台用ascii转义，避免编码问题
        logger.error(f"{message}: {ascii(str(exception))}")
        # 文件记录完整堆栈
        logger.bind(file_only=True).exception(f"{message}")
    else:
        logger.error(message)


class TaskManager:
    """任务管理器"""

    TASK_TIMEOUT_SECONDS = 600  # 任务超时时间：10分钟
    CLIENT_COOLDOWN_SECONDS = 3  # 客户端冷却时间：3秒

    def __init__(self):
        self.tasks = []
        self.current_index = 0
        self.is_running = False
        self.clients = {}
        self.next_page_number = 1

    def register_client(self, websocket, page_url):
        import time
        for cid, info in list(self.clients.items()):
            if info['url'] == page_url:
                del self.clients[cid]

        client_id = f"c{len(self.clients)}_{int(time.time()) % 10000}"
        page_number = self.next_page_number
        self.next_page_number += 1
        self.clients[client_id] = {
            'ws': websocket,
            'url': page_url,
            'busy': False,
            'task_id': None,
            'page_number': page_number,
            'last_task_end': None
        }
        logger.info(f"客户端注册: {client_id} (页面{page_number})")
        return client_id, page_number

    def remove_client(self, client_id):
        if client_id in self.clients:
            page_number = self.clients[client_id].get('page_number')
            task_id = self.clients[client_id]['task_id']
            if task_id:
                for task in self.tasks:
                    if task['id'] == task_id and task['status'] == '处理中':
                        task['status'] = '等待中'
                        logger.warning(f"任务 {task_id} 因客户端断开重置为等待")
            del self.clients[client_id]
            logger.info(f"客户端断开: {client_id} (页面{page_number})")

    def get_idle_client(self):
        now = datetime.now()
        for cid, info in self.clients.items():
            if not info['busy']:
                # 检查冷却时间
                last_end = info.get('last_task_end')
                if last_end:
                    elapsed = (now - datetime.fromisoformat(last_end)).total_seconds()
                    if elapsed < self.CLIENT_COOLDOWN_SECONDS:
                        continue  # 还在冷却中，跳过这个客户端
                return cid, info
        return None, None

    def mark_client_busy(self, client_id, task_id):
        if client_id in self.clients:
            self.clients[client_id]['busy'] = True
            self.clients[client_id]['task_id'] = task_id
            for task in self.tasks:
                if task['id'] == task_id:
                    task['client_id'] = client_id
                    break

    def mark_client_idle(self, client_id):
        if client_id in self.clients:
            self.clients[client_id]['busy'] = False
            self.clients[client_id]['task_id'] = None
            self.clients[client_id]['last_task_end'] = datetime.now().isoformat()

    def get_client_count(self):
        total = len(self.clients)
        busy = sum(1 for c in self.clients.values() if c['busy'])
        return total, busy

    def update_task_status_detail(self, task_id, status_detail):
        for task in self.tasks:
            if task['id'] == task_id:
                task['status_detail'] = status_detail
                return True
        return False

    def add_task(self, prompt, task_type, aspect_ratio, resolution,
                 reference_images=None, output_dir=None, import_row_number=None):
        prompt = prompt.strip()
        if not prompt:
            return None

        if task_type == "Text to Video":
            reference_images = []

        task_id = f"task_{len(self.tasks)}_{datetime.now().strftime('%H%M%S%f')}"
        file_ext = ".mp4" if "Video" in task_type else ".png"

        task = {
            'id': task_id,
            'prompt': prompt,
            'status': '等待中',
            'status_detail': '',
            'file_ext': file_ext,
            'output_dir': output_dir,
            'client_id': None,
            'task_type': task_type,
            'aspect_ratio': aspect_ratio,
            'resolution': resolution,
            'reference_images': reference_images or [],
            'start_time': None,
            'end_time': None,
            'import_row_number': import_row_number  # 导入任务的行号（编号）
        }
        self.tasks.append(task)
        logger.info(f"添加任务: {task_id} | {task_type} | {aspect_ratio}")
        return task

    def get_next_task(self):
        while self.current_index < len(self.tasks):
            task = self.tasks[self.current_index]
            if task['status'] == '等待中':
                return task
            self.current_index += 1
        return None

    def check_timeout_tasks(self):
        """检查并处理超时任务，返回超时的任务列表"""
        timeout_tasks = []
        now = datetime.now()
        for task in self.tasks:
            if task['status'] == '处理中' and task.get('start_time'):
                start = datetime.fromisoformat(task['start_time'])
                elapsed = (now - start).total_seconds()
                if elapsed > self.TASK_TIMEOUT_SECONDS:
                    task['status'] = '超时'
                    task['end_time'] = now.isoformat()
                    task['status_detail'] = f'任务超时（超过{self.TASK_TIMEOUT_SECONDS // 60}分钟）'
                    logger.warning(f"任务超时: {task['id']} (耗时 {elapsed:.0f}s)")
                    # 释放对应客户端
                    client_id = task.get('client_id')
                    if client_id and client_id in self.clients:
                        self.clients[client_id]['busy'] = False
                        self.clients[client_id]['task_id'] = None
                    timeout_tasks.append(task)
        return timeout_tasks


class ImageProcessor:
    @staticmethod
    def compress_image_to_base64(image_path, max_size_bytes=128 * 1024):
        try:
            img = Image.open(image_path)
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            quality = 95
            while quality > 5:
                buffer = io.BytesIO()
                img.save(buffer, format='JPEG', quality=quality, optimize=True)
                size = buffer.tell()
                if size <= max_size_bytes:
                    buffer.seek(0)
                    return base64.b64encode(buffer.getvalue()).decode('utf-8')
                quality -= 5

            scale = 0.9
            while scale > 0.1:
                new_size = (int(img.size[0] * scale), int(img.size[1] * scale))
                resized_img = img.resize(new_size, Image.Resampling.LANCZOS)
                buffer = io.BytesIO()
                resized_img.save(buffer, format='JPEG', quality=85, optimize=True)
                size = buffer.tell()
                if size <= max_size_bytes:
                    buffer.seek(0)
                    return base64.b64encode(buffer.getvalue()).decode('utf-8')
                scale -= 0.1

            buffer.seek(0)
            return base64.b64encode(buffer.getvalue()).decode('utf-8')
        except Exception as e:
            log_error_to_file("压缩图片失败", e)
            return None

    @staticmethod
    def generate_thumbnail(image_path, size=(200, 200)):
        """生成缩略图，返回 base64"""
        try:
            img = Image.open(image_path)
            # 转换为 RGB（处理 RGBA 等）
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # 生成缩略图
            img.thumbnail(size, Image.Resampling.LANCZOS)
            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=85, optimize=True)
            buffer.seek(0)
            return base64.b64encode(buffer.read()).decode('utf-8')
        except Exception as e:
            log_error_to_file("生成缩略图失败", e)
            return None


class ImageDownloader:
    @staticmethod
    async def save_base64_image(base64_data, filename, output_dir=None):
        if output_dir is None:
            output_dir = OUTPUT_DIR
        filepath = Path(output_dir) / filename
        try:
            image_data = base64.b64decode(base64_data)
            with open(filepath, 'wb') as f:
                f.write(image_data)
            # 返回绝对路径
            return filepath.absolute()
        except Exception as e:
            log_error_to_file("保存图片失败", e)
            return None


class GuideServer:
    """引导页面 HTTP 服务器"""

    def __init__(self, port=12346):
        self.port = port
        self.server = None
        self.thread = None

    def start(self):
        """启动 HTTP 服务器"""
        if getattr(sys, 'frozen', False):
            guide_dir = Path(sys._MEIPASS) / 'guide'
        else:
            guide_dir = Path(__file__).parent / 'guide'

        if not guide_dir.exists():
            logger.warning(f"引导页面目录不存在: {guide_dir}")
            return False

        try:
            handler = partial(SimpleHTTPRequestHandler, directory=str(guide_dir))
            self.server = HTTPServer(('localhost', self.port), handler)
            self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
            self.thread.start()
            logger.info(f"引导页面服务已启动: http://localhost:{self.port}")
            return True
        except OSError as e:
            logger.error(f"启动引导页面服务失败: {e}")
            return False

    def stop(self):
        """停止 HTTP 服务器"""
        if self.server:
            self.server.shutdown()
            logger.info("引导页面服务已停止")


class WebSocketServer:
    def __init__(self, task_manager):
        self.task_manager = task_manager
        self.server = None
        self.chunk_buffer = {}

    def log(self, message):
        logger.info(message)

    async def handler(self, websocket):
        client_id = None
        page_number = None
        try:
            first_msg = await websocket.recv()
            data = json.loads(first_msg)

            if data.get('type') != 'register':
                self.log(f"[警告] 首条消息不是注册消息，断开连接")
                return

            page_url = data.get('page_url', 'unknown')
            client_id, page_number = self.task_manager.register_client(websocket, page_url)
            total, busy = self.task_manager.get_client_count()
            self.log(f"[OK] 客户端注册: {client_id} (页面#{page_number})，当前连接数: {total}")

            await websocket.send(json.dumps({
                'type': 'register_success',
                'client_id': client_id
            }))

            async for message in websocket:
                data = json.loads(message)
                msg_type = data.get("type")

                if msg_type == "image_chunk":
                    task_id = data.get("task_id")
                    chunk_index = data.get("chunk_index")
                    total_chunks = data.get("total_chunks")
                    chunk_data = data.get("data")

                    if task_id not in self.chunk_buffer:
                        self.chunk_buffer[task_id] = {}

                    self.chunk_buffer[task_id][chunk_index] = chunk_data
                    self.log(f"[收到] [#{page_number}] 收到分块 {chunk_index + 1}/{total_chunks}")

                    if len(self.chunk_buffer[task_id]) == total_chunks:
                        full_base64 = ''.join(
                            self.chunk_buffer[task_id][i]
                            for i in range(total_chunks)
                        )
                        del self.chunk_buffer[task_id]
                        self.log(f"[OK] [#{page_number}] 分块合并完成，总大小: {len(full_base64) // 1024} KB")
                        await self.handle_image_result(client_id, task_id, full_base64)

                elif msg_type == "image_data":
                    task_id = data.get("task_id")
                    image_data = data.get("data")
                    self.log(f"[收到] [#{page_number}] 收到图片数据，大小: {len(image_data) // 1024} KB")
                    await self.handle_image_result(client_id, task_id, image_data)

                elif msg_type == "result":
                    task_id = data.get("task_id")
                    error = data.get("error")
                    if error:
                        self.log(f"[失败] [#{page_number}] 任务失败: {error}")
                        for task in self.task_manager.tasks:
                            if task['id'] == task_id:
                                task['status'] = '失败'
                                task['status_detail'] = error
                                task['end_time'] = datetime.now().isoformat()
                                break
                    self.task_manager.mark_client_idle(client_id)

                elif msg_type == "status":
                    status_msg = data.get('message', '')
                    self.log(f"[状态] [#{page_number}] {status_msg}")
                    task_id = self.task_manager.clients.get(client_id, {}).get('task_id')
                    if task_id:
                        self.task_manager.update_task_status_detail(task_id, status_msg)

        except Exception as e:
            self.log("连接异常")
            log_error_to_file("WebSocket连接异常", e)
        finally:
            if client_id:
                self.task_manager.remove_client(client_id)
                total, busy = self.task_manager.get_client_count()
                self.log(f"[断开] 客户端断开: {client_id} (页面#{page_number})，当前连接数: {total}")

    async def handle_image_result(self, client_id, task_id, base64_data):
        for task in self.task_manager.tasks:
            if task['id'] == task_id:
                output_dir = task.get('output_dir')
                if output_dir:
                    if not Path(output_dir).is_absolute():
                        output_dir = OUTPUT_DIR / output_dir
                    else:
                        output_dir = Path(output_dir)
                else:
                    output_dir = OUTPUT_DIR
                output_dir.mkdir(parents=True, exist_ok=True)

                file_ext = task.get('file_ext', '.png')

                # 如果是导入任务，使用行号作为文件名（不论是否设置输出文件夹）
                if task.get('import_row_number'):
                    filename = f"{task['import_row_number']}{file_ext}"
                    filepath = output_dir / filename
                else:
                    # 普通任务使用时间戳
                    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                    filename = f"{timestamp}{file_ext}"
                    filepath = output_dir / filename

                    # 避免重名
                    counter = 1
                    while filepath.exists():
                        filename = f"{timestamp}_{counter}{file_ext}"
                        filepath = output_dir / filename
                        counter += 1

                # 保存文件
                saved = await ImageDownloader.save_base64_image(base64_data, filename, output_dir)
                if saved:
                    task['status'] = '已完成'
                    task['end_time'] = datetime.now().isoformat()
                    task['saved_path'] = str(saved)
                    task['output_dir_path'] = str(output_dir)
                    # 生成并缓存缩略图（只对图片生成）
                    if task.get('file_ext') in ('.png', '.jpg'):
                        try:
                            task['preview_base64'] = ImageProcessor.generate_thumbnail(str(saved), size=(200, 200))
                        except Exception:
                            task['preview_base64'] = ''
                    logger.info(f"任务完成: {task_id} -> {saved}")
                else:
                    task['status'] = '下载失败'
                    task['end_time'] = datetime.now().isoformat()
                    logger.error(f"任务保存失败: {task_id}")
                break
        self.task_manager.mark_client_idle(client_id)

    async def start(self):
        try:
            self.server = await serve(
                self.handler,
                "localhost",
                12345,
                max_size=50 * 1024 * 1024
            )
            logger.info("WebSocket 服务器已启动: ws://localhost:12345")
        except OSError as e:
            # 端口被占用
            logger.error("WebSocket 服务器启动失败 (端口占用)")
            log_error_to_file("WebSocket服务器启动失败", e)
            raise

    async def stop(self):
        if self.server:
            self.server.close()
            await self.server.wait_closed()


class Api:
    """暴露给前端的 API"""

    def __init__(self, task_manager, loop):
        self.task_manager = task_manager
        self.loop = loop

    def add_task(self, prompt, task_type, aspect_ratio, resolution, reference_images, output_dir):
        task = self.task_manager.add_task(
            prompt=prompt,
            task_type=task_type,
            aspect_ratio=aspect_ratio,
            resolution=resolution,
            reference_images=reference_images or [],
            output_dir=output_dir or None
        )
        if task:
            logger.info(f"已添加任务，当前共 {len(self.task_manager.tasks)} 个")
            return {'success': True}
        return {'success': False, 'error': '添加失败'}

    def get_status(self):
        total, busy = self.task_manager.get_client_count()
        tasks_data = []
        for t in self.task_manager.tasks:
            tasks_data.append({
                'id': t['id'],
                'prompt': t['prompt'],
                'status': t['status'],
                'status_detail': t.get('status_detail', ''),
                'task_type': t['task_type'],
                'aspect_ratio': t['aspect_ratio'],
                'resolution': t['resolution'],
                'saved_path': t.get('saved_path', ''),
                'output_dir': t.get('output_dir', ''),
                'start_time': t.get('start_time'),
                'end_time': t.get('end_time'),
                'file_ext': t.get('file_ext', ''),
                'preview_base64': t.get('preview_base64', '')
            })
        return {
            'client_count': total,
            'busy_count': busy,
            'is_running': self.task_manager.is_running,
            'tasks': tasks_data
        }

    def start_execution(self):
        total, _ = self.task_manager.get_client_count()
        if total == 0:
            logger.warning("启动执行失败: 没有连接的客户端")
            return
        if not self.task_manager.tasks:
            logger.warning("启动执行失败: 任务列表为空")
            return
        self.task_manager.is_running = True
        logger.info(f"启动任务执行: 客户端数={total}, 任务数={len(self.task_manager.tasks)}")
        asyncio.run_coroutine_threadsafe(self._execute_tasks(), self.loop)

    def stop_execution(self):
        self.task_manager.is_running = False
        logger.info("已停止执行")

    async def _execute_tasks(self):
        logger.info("任务执行循环启动")

        while self.task_manager.is_running:
            # 检查超时任务
            timeout_tasks = self.task_manager.check_timeout_tasks()
            for t in timeout_tasks:
                logger.warning(f"任务超时: {t['id']}")

            task = self.task_manager.get_next_task()
            if not task:
                has_busy = any(c['busy'] for c in self.task_manager.clients.values())
                if not has_busy:
                    logger.info("所有任务已完成")
                    break
                await asyncio.sleep(1)
                continue

            client_id, client_info = self.task_manager.get_idle_client()
            if not client_info:
                await asyncio.sleep(1)
                continue

            # 检查导入任务的文件是否已存在，如果存在直接跳过
            if task.get('import_row_number'):
                output_dir = task.get('output_dir')
                if output_dir:
                    if not Path(output_dir).is_absolute():
                        output_dir = OUTPUT_DIR / output_dir
                    else:
                        output_dir = Path(output_dir)
                else:
                    output_dir = OUTPUT_DIR

                file_ext = task.get('file_ext', '.png')
                filepath = output_dir / f"{task['import_row_number']}{file_ext}"

                if filepath.exists():
                    task['status'] = '已完成'
                    task['end_time'] = datetime.now().isoformat()
                    task['saved_path'] = str(filepath.absolute())
                    task['output_dir_path'] = str(output_dir)
                    task['status_detail'] = '文件已存在，跳过生成'
                    # 生成缩略图
                    if file_ext in ('.png', '.jpg'):
                        try:
                            task['preview_base64'] = ImageProcessor.generate_thumbnail(str(filepath), size=(200, 200))
                        except Exception:
                            task['preview_base64'] = ''
                    logger.info(f"[跳过] 文件已存在: {filepath}")
                    self.task_manager.current_index += 1
                    continue

            task['status'] = '处理中'
            task['start_time'] = datetime.now().isoformat()
            self.task_manager.mark_client_busy(client_id, task['id'])
            self.task_manager.current_index += 1

            logger.info(f"分配任务: {task['id']} -> {client_id} | {task['task_type']}")

            # 延迟处理参考图片：如果是路径则压缩为 base64
            reference_images = []
            for img in task['reference_images']:
                if img and not img.startswith('/9j/') and not img.startswith('iVBOR') and Path(img).exists():
                    # 是文件路径，需要压缩
                    base64_data = ImageProcessor.compress_image_to_base64(img)
                    if base64_data:
                        reference_images.append(base64_data)
                else:
                    # 已经是 base64 数据
                    reference_images.append(img)

            message = json.dumps({
                'type': 'task',
                'task_id': task['id'],
                'prompt': task['prompt'],
                'task_type': task['task_type'],
                'aspect_ratio': task['aspect_ratio'],
                'resolution': task['resolution'],
                'reference_images': reference_images
            })

            try:
                await client_info['ws'].send(message)
            except Exception as e:
                logger.error(f"任务发送失败: {task['id']} -> {client_id}")
                log_error_to_file(f"任务发送失败", e)
                task['status'] = '等待中'
                self.task_manager.mark_client_idle(client_id)

            await asyncio.sleep(0.5)

        self.task_manager.is_running = False
        logger.info("任务队列执行结束")

    def select_images(self):
        """打开文件对话框选择图片"""
        file_types = ('图片文件 (*.png;*.jpg;*.jpeg;*.gif;*.bmp;*.webp)',)
        result = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,
            allow_multiple=True,
            file_types=file_types
        )
        if not result:
            return []

        images = []
        for filepath in result:
            logger.info(f"正在处理: {Path(filepath).name}")
            base64_data = ImageProcessor.compress_image_to_base64(filepath)
            if base64_data:
                images.append(base64_data)
                size_kb = len(base64_data) * 3 / 4 / 1024
                logger.info(f"已添加: {Path(filepath).name} (压缩后 ~{size_kb:.1f}KB)")
        return images

    def import_excel(self):
        """导入 Excel 文件"""
        if load_workbook is None:
            return {'success': False, 'count': 0, 'errors': ['请安装 openpyxl']}

        file_types = ('Excel文件 (*.xlsx)',)
        result = webview.windows[0].create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=file_types
        )
        if not result:
            return {'success': False, 'count': 0, 'errors': []}

        filepath = result[0]
        logger.info(f"导入 Excel 开始: {Path(filepath).name}")

        # 验证分辨率和任务类型是否匹配
        def validate_resolution(task_type, resolution, aspect_ratio):
            """检查分辨率是否与任务类型兼容"""
            valid_resolutions = {
                "Create Image": ["4K", "2K", "1K"],
                "Text to Video": ["1080p", "720p"],
                "Frames to Video": ["1080p", "720p"],
                "Ingredients to Video": ["1080p", "720p"]
            }
            if task_type not in valid_resolutions:
                return False, f"未知任务类型: {task_type}"

            allowed = valid_resolutions[task_type]

            # 视频类任务竖屏时不支持1080p
            if "Video" in task_type and aspect_ratio == "9:16" and resolution == "1080p":
                return False, f"{task_type} 竖屏模式不支持 1080p，请使用 720p"

            if resolution not in allowed:
                return False, f"{task_type} 不支持分辨率 {resolution}，请使用: {', '.join(allowed)}"
            return True, ""

        task_type_map = {
            "文生图片": "Create Image",
            "文生视频": "Text to Video",
            "图生视频": "Ingredients to Video",
            "首尾帧视频": "Frames to Video",
        }
        orientation_map = {
            "横屏": "16:9",
            "竖屏": "9:16"
        }

        try:
            wb = load_workbook(filepath)
            ws = wb.active

            # 第一步：验证所有行
            tasks_to_add = []
            validation_errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[1]:
                    continue

                try:
                    # 读取编号列（第一列）
                    row_number = str(row[0]).strip() if row[0] else str(row_idx)

                    prompt = str(row[1]).strip() if row[1] else ""
                    if not prompt:
                        continue

                    task_type_cn = str(row[2]).strip() if len(row) > 2 and row[2] else "图片"
                    orientation_cn = str(row[3]).strip() if len(row) > 3 and row[3] else "横屏"
                    resolution = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    output_dir = str(row[5]).strip() if len(row) > 5 and row[5] else None

                    # 验证任务类型
                    if task_type_cn not in task_type_map:
                        validation_errors.append(f"编号{row_number}: 未知任务类型: {task_type_cn}，请使用: {', '.join(task_type_map.keys())}")
                        continue

                    task_type = task_type_map[task_type_cn]
                    aspect_ratio = orientation_map.get(orientation_cn, "16:9")

                    if not resolution:
                        resolution = "1080p" if "Video" in task_type else "4K"
                    else:
                        # 分辨率忽略大小写处理
                        resolution_upper = resolution.upper()
                        resolution_lower = resolution.lower()
                        # 标准化为正确的格式（如 4k -> 4K, 1080p -> 1080p）
                        if resolution_upper in ["4K", "2K", "1K"]:
                            resolution = resolution_upper
                        elif resolution_lower == "1080p":
                            resolution = "1080p"
                        elif resolution_lower == "720p":
                            resolution = "720p"

                    # 验证分辨率
                    is_valid, error_msg = validate_resolution(task_type, resolution, aspect_ratio)
                    if not is_valid:
                        validation_errors.append(f"行{row_idx}: {error_msg}")
                        continue

                    reference_images = []
                    max_images = {
                        "Create Image": 8,
                        "Frames to Video": 2,
                        "Ingredients to Video": 3,
                        "Text to Video": 0
                    }.get(task_type, 8)

                    # 只收集图片路径，不在导入时压缩（延迟到执行时处理）
                    for i in range(max_images):
                        col_idx = 6 + i
                        if len(row) > col_idx and row[col_idx]:
                            img_path = str(row[col_idx]).strip()
                            if img_path and Path(img_path).exists():
                                reference_images.append(img_path)

                    tasks_to_add.append({
                        'prompt': prompt,
                        'task_type': task_type,
                        'aspect_ratio': aspect_ratio,
                        'resolution': resolution,
                        'reference_images': reference_images,
                        'output_dir': output_dir,
                        'import_row_number': row_number  # Excel 编号列的值
                    })

                except Exception as e:
                    validation_errors.append(f"编号{row_number}: {str(e)}")

            wb.close()

            # 如果有验证错误，全部不导入
            if validation_errors:
                return {'success': False, 'count': 0, 'errors': validation_errors}

            # 如果没有任何有效任务
            if not tasks_to_add:
                return {'success': False, 'count': 0, 'errors': ['没有找到有效的任务行']}

            # 第二步：全部验证通过，一次性导入所有任务
            for task_data in tasks_to_add:
                self.task_manager.add_task(
                    prompt=task_data['prompt'],
                    task_type=task_data['task_type'],
                    aspect_ratio=task_data['aspect_ratio'],
                    resolution=task_data['resolution'],
                    reference_images=task_data['reference_images'],
                    output_dir=task_data['output_dir'],
                    import_row_number=task_data['import_row_number']
                )

            count = len(tasks_to_add)
            logger.info(f"从Excel导入 {count} 个任务")
            return {'success': True, 'count': count, 'errors': []}

        except Exception as e:
            return {'success': False, 'count': 0, 'errors': [str(e)]}

    def export_template(self):
        """导出 Excel 模板"""
        if Workbook is None:
            return

        file_types = ('Excel文件 (*.xlsx)',)
        result = webview.windows[0].create_file_dialog(
            webview.SAVE_DIALOG,
            file_types=file_types,
            save_filename='任务模板.xlsx'
        )
        if not result:
            return

        filepath = result if isinstance(result, str) else result[0]

        # 确保有 .xlsx 扩展名
        if not filepath.lower().endswith('.xlsx'):
            filepath += '.xlsx'

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"

            headers = ["编号", "提示词", "任务类型", "屏幕方向", "分辨率", "输出文件夹",
                       "图1", "图2", "图3", "图4", "图5", "图6", "图7", "图8"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            examples = [
                [1, "A beautiful sunset over the ocean", "文生图片", "横屏", "4K", "sunset"],
                [2, "A beautiful moon over the ocean", "文生图片", "竖屏", "2K", "sunset"],
                [3, "A cute cat playing", "文生视频", "横屏", "1080p", "cats"],
                [4, "A cute dog playing", "文生视频", "竖屏", "720p", "dogs_注意veo3竖屏视频不支持1080p"],
                [5, "动起来", "首尾帧视频", "横屏", "1080p", "frames", "/Users/wei/Downloads/pig.jpeg"],
                [6, "组合这些照片为一个创意视频", "图生视频", "横屏", "1080p", "collage", "/Users/wei/Downloads/pig.jpeg"],
            ]

            for row_idx, example in enumerate(examples, start=2):
                for col_idx, value in enumerate(example, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(filepath)
            logger.info(f"已导出模板: {filepath}")

        except Exception as e:
            log_error_to_file("导出模板失败", e)

    def open_output_dir(self):
        """打开输出目录"""
        self._open_directory(OUTPUT_DIR)

    def open_logs_dir(self):
        """打开日志目录"""
        self._open_directory(LOGS_DIR)

    def open_task_file(self, task_index):
        """打开任务的文件（不是文件夹）"""
        if 0 <= task_index < len(self.task_manager.tasks):
            task = self.task_manager.tasks[task_index]
            saved_path = task.get('saved_path')
            if saved_path and Path(saved_path).exists():
                path = Path(saved_path)
                # 直接打开文件
                path_str = str(path.absolute())
                system = platform.system()

                if system == 'Windows':
                    os.startfile(path_str)
                elif system == 'Darwin':
                    subprocess.run(['open', path_str])
                else:
                    subprocess.run(['xdg-open', path_str])
            else:
                # 文件不存在，打开所在目录
                self.open_task_dir(task_index)
        else:
            self.open_output_dir()

    def open_task_dir(self, task_index):
        """打开任务的输出目录"""
        if 0 <= task_index < len(self.task_manager.tasks):
            task = self.task_manager.tasks[task_index]
            output_dir = task.get('output_dir_path', str(OUTPUT_DIR))
            self._open_directory(Path(output_dir))
        else:
            self._open_directory(OUTPUT_DIR)

    def _open_directory(self, path):
        path = Path(path)
        if not path.exists():
            path.mkdir(parents=True, exist_ok=True)

        path_str = str(path.absolute())
        system = platform.system()

        if system == 'Windows':
            os.startfile(path_str)
        elif system == 'Darwin':
            subprocess.run(['open', path_str])
        else:
            subprocess.run(['xdg-open', path_str])

    def get_app_version(self) -> str:
        """获取应用版本号"""
        version = get_version()
        logger.info(f"获取应用版本: {version}")
        return version

    def check_update(self) -> dict:
        """检查更新"""
        logger.info("前端请求检查更新")
        info = check_for_updates()

        if info is None:
            logger.warning("更新检查返回 None，检查失败")
            return {
                'success': False,
                'has_update': False,
                'current_version': get_version(),
                'latest_version': '',
                'release_notes': '',
                'download_url': '',
                'release_url': ''
            }

        logger.info(f"更新检查完成: 有更新={info.has_update}, 最新版本={info.latest_version}")
        return {
            'success': True,
            'has_update': info.has_update,
            'current_version': info.current_version,
            'latest_version': info.latest_version,
            'release_notes': info.release_notes,
            'download_url': info.download_url,
            'release_url': info.release_url
        }

    def open_update_page(self, url: str) -> bool:
        """在浏览器中打开更新下载页面"""
        logger.info(f"打开下载页面: {url}")
        return open_download_page(url)

    def open_guide_page(self) -> bool:
        """在外部浏览器中打开引导页面"""
        guide_url = "http://localhost:12346"
        logger.info(f"打开引导页面: {guide_url}")
        try:
            webbrowser.open(guide_url)
            return True
        except Exception as e:
            logger.error(f"打开引导页面失败: {e}")
            return False


def run_async_loop(loop):
    asyncio.set_event_loop(loop)
    loop.run_forever()


def main():
    # 启动日志
    logger.info("=" * 50)
    logger.info(f"Veo3Free 启动 - 版本: {get_version()}")
    logger.info(f"运行环境: {'打包模式' if getattr(sys, 'frozen', False) else '开发模式'}")
    logger.info(f"操作系统: {platform.system()} {platform.release()}")
    logger.info(f"Python: {sys.version}")
    logger.info(f"输出目录: {OUTPUT_DIR}")
    logger.info(f"日志目录: {LOGS_DIR}")
    logger.info("=" * 50)

    # 创建事件循环
    loop = asyncio.new_event_loop()
    thread = threading.Thread(target=run_async_loop, args=(loop,), daemon=True)
    thread.start()

    # 创建任务管理器和 API
    task_manager = TaskManager()
    api = Api(task_manager, loop)

    # 启动 WebSocket 服务器，捕获端口占用错误
    ws_server = WebSocketServer(task_manager)
    logger.info("正在启动 WebSocket 服务器 (端口 12345)...")
    ws_start_future = asyncio.run_coroutine_threadsafe(ws_server.start(), loop)

    try:
        # 等待 WebSocket 启动完成（超时 5 秒）
        ws_start_future.result(timeout=5)
        logger.info("WebSocket 服务器启动成功")
    except OSError as e:
        # 端口被占用，弹框提示用户
        logger.error(f"WebSocket 服务器启动失败 (端口占用): {e}")
        error_msg = "无法启动应用!\n\nWebSocket 端口 12345 被占用\n\n请检查是否有其他程序占用该端口，或稍后重试。"
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口
        messagebox.showerror("启动失败", error_msg)
        root.destroy()
        return
    except Exception as e:
        # 其他错误
        logger.error(f"WebSocket 服务器启动失败: {e}")
        error_msg = "无法启动 WebSocket 服务器，请稍后重试。"
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("启动失败", error_msg)
        root.destroy()
        return

    # 启动引导页面服务
    logger.info("正在启动引导页面服务 (端口 12346)...")
    guide_server = GuideServer(port=12346)
    if guide_server.start():
        logger.info("引导页面服务启动成功")
    else:
        logger.warning("引导页面服务启动失败")

    # 确定 web 目录和 URL
    if getattr(sys, 'frozen', False):
        # 打包后，使用打包的web目录
        web_dir = Path(sys._MEIPASS) / 'web'
        url = str(web_dir / 'index.html')
        logger.info(f"使用打包资源: {url}")
    else:
        # 开发模式
        web_dir = Path(__file__).parent / 'web'
        # 检查是否使用开发服务器
        if os.environ.get('DEV') == '1' or not web_dir.exists():
            url = 'http://localhost:5173'
            logger.info("使用开发服务器: http://localhost:5173")
        else:
            url = str(web_dir / 'index.html')
            logger.info(f"使用本地文件: {url}")

    # 创建窗口
    logger.info("正在创建应用窗口...")
    window = webview.create_window(
        'Veo3Free - AI生成工具',
        url,
        width=1000,
        height=700,
        min_size=(800, 600),
        maximized=True,
        js_api=api
    )
    # !!! 严谨对api设置window等对象，例如"api.window = window"是极其危险的！！！

    # 启动 webview
    logger.info("启动 webview 主循环...")
    webview.start()

    # 清理
    logger.info("正在关闭应用...")
    guide_server.stop()
    asyncio.run_coroutine_threadsafe(ws_server.stop(), loop)
    loop.call_soon_threadsafe(loop.stop)
    logger.info("应用已退出")


if __name__ == "__main__":
    main()
